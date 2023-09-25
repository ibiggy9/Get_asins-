import openpyxl

def american():
    # Load the workbook
    wb = openpyxl.load_workbook('YourASINDownloadxlsx')
    
    # Select the specific worksheet for US
    ws = wb['US_Detail View_Sales Diagnostic']
    
    # Extract all the values from the first column of the worksheet into mylist
    mylist = []
    for col_cells in ws.iter_cols(min_col=1, max_col=1):
        for cell in col_cells:    
            mylist.append(cell.value)

    # Filter out the values that contain the condition 'B0' and store them in US_ASINS list
    US_ASINS = []
    condition = 'B0'
    for i in mylist:
        if condition in i:
            US_ASINS.append(i)

    # Construct the complete Amazon URL for each ASIN in the US_ASINS list
    url_default = 'https://www.amazon.com/dp/'
    url_list = []
    for i in US_ASINS:
        url_list.append(f'{url_default}{i}')
    
    # Return the list of URLs
    return url_list

def canadian():
    # Load the workbook
    wb = openpyxl.load_workbook('yourASINDownload.xlsx')
    
    # Select the specific worksheet for Canada
    ws = wb['CA_Detail View_Sales Diagnostic']
    
    # Extract all the values from the first column of the worksheet into mylist
    mylist = []
    for col_cells in ws.iter_cols(min_col=1, max_col=1):
        for cell in col_cells:    
            mylist.append(cell.value)

    # Filter out the values that contain the condition 'B0' and store them in CA_ASINS list
    CA_ASINS = []
    condition = 'B0'
    for i in mylist:
        if condition in i:
            CA_ASINS.append(i)
            
    # Construct the complete Amazon URL for each ASIN in the CA_ASINS list
    url_default = 'https://www.amazon.ca/dp/'
    url_list = []
    for i in CA_ASINS:
        url_list.append(f'{url_default}{i}')
    
    # Return the list of URLs
    return url_list
